"""
StatementIQ – FastAPI Backend
Wraps Abakkus PMS + Kotak 811 (Bank / Credit Card) parsers.

Run:
    cd backend
    uvicorn main:app --reload --host 0.0.0.0 --port 8000

Install once:
    pip install fastapi uvicorn[standard] pdfplumber pandas openpyxl watchdog \
                python-dateutil pytesseract Pillow
    # For Kotak CC OCR also install Ghostscript 10.x and Tesseract 5.x
"""

import io
import base64
import re
import shutil
import os
import tempfile
import subprocess
import asyncio
import signal
import gc
from datetime import datetime
from pathlib import Path
from typing import Optional
import xml.etree.ElementTree as ET

import pandas as pd
import pdfplumber
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles

# ── Optional OCR imports ──────────────────────────────────────────────────────
try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# ── PDF splitting for large files ─────────────────────────────────────────────
try:
    import pikepdf
    PIKEPDF_AVAILABLE = True
except ImportError:
    PIKEPDF_AVAILABLE = False

# ── Config ────────────────────────────────────────────────────────────────────
import platform

if platform.system() == "Windows":
    TESSERACT_EXE = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    GS_EXE        = r"C:\Program Files\gs\gs10.06.0\bin\gswin64c.exe"
else:
    # Linux / macOS (Render, Docker, etc.) – resolve via PATH
    TESSERACT_EXE = shutil.which("tesseract") or "tesseract"
    GS_EXE        = shutil.which("gs") or "gs"

OCR_DPI       = 600
OCR_DPI_LARGE_DOC = 600
OCR_HIGH_PAGE_COUNT = 12
GS_TIMEOUT_BASE_SECS = 180
MIN_ROWS      = 1
OCR_DATE_HINT_RE = re.compile(r"\b\d{1,2}[\-/\.\s]\d{1,2}[\-/\.\s]\d{2,4}\b")
OCR_AMOUNT_HINT_RE = re.compile(r"\b\d{1,3}(?:,\d{3})*\.\d{2}\b")

INPUT_DIR  = Path(__file__).parent / "input"
OUTPUT_DIR = Path(__file__).parent / "output"
INPUT_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Processing limits ─────────────────────────────────────────────────────────
MAX_PDF_SIZE_MB = 100         # reject files larger than this
CONVERT_TIMEOUT_SECS = 540   # timeout for /convert processing (9 min — upgraded Render plan)
LARGE_PDF_PAGE_THRESHOLD = 15  # warn above this many pages

app = FastAPI(title="StatementIQ API", version="1.0.0")

ALLOWED_ORIGINS = os.environ.get("CORS_ORIGINS", "").split(",") if os.environ.get("CORS_ORIGINS") else ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── JWT / Auth helpers ────────────────────────────────────────────────────────
from fastapi import Request

def _env_clean(name: str, default: str = "") -> str:
    value = os.environ.get(name, default)
    if not isinstance(value, str):
        return default
    return value.strip().strip('"').strip("'").strip() or default


SUPABASE_JWT_SECRET = _env_clean("SUPABASE_JWT_SECRET")
SUPABASE_URL = _env_clean(
    "SUPABASE_URL",
    _env_clean("VITE_SUPABASE_URL", "https://tfkqovfodkstoqjhqegv.supabase.co"),
)
SUPABASE_PUBLISHABLE_KEY = _env_clean(
    "SUPABASE_PUBLISHABLE_KEY",
    _env_clean(
        "VITE_SUPABASE_PUBLISHABLE_KEY",
        "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRma3FvdmZvZGtzdG9xamhxZWd2Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzQwOTYzNzMsImV4cCI6MjA4OTY3MjM3M30.xIHyzzsaz4wnw9KvRycc-I9lOhoYvtgID9Id7Qj4uZw",
    ),
)
SUPABASE_SERVICE_ROLE_KEY = _env_clean("SUPABASE_SERVICE_ROLE_KEY", _env_clean("VITE_SUPABASE_SERVICE_ROLE_KEY"))

def _get_current_user(request: Request) -> dict:
    """Validate JWT and return user claims. Raises HTTPException on failure."""
    auth_header = request.headers.get("authorization", "")
    parts = auth_header.split(" ", 1)
    if len(parts) != 2 or parts[0].lower() != "bearer" or not parts[1].strip():
        raise HTTPException(401, "Missing or invalid Authorization header")
    token = parts[1].strip()

    # 1) Primary verification via Auth API (resilient to key/algorithm mismatch)
    verify_key = SUPABASE_PUBLISHABLE_KEY or SUPABASE_SERVICE_ROLE_KEY
    if SUPABASE_URL and verify_key:
        import json as _json
        import urllib.error
        import urllib.request

        verify_url = f"{SUPABASE_URL}/auth/v1/user"
        verify_req = urllib.request.Request(
            verify_url,
            headers={
                "apikey": verify_key,
                "Authorization": f"Bearer {token}",
                "Accept": "application/json",
            },
            method="GET",
        )
        try:
            with urllib.request.urlopen(verify_req, timeout=8) as resp:
                user = _json.loads(resp.read())
                user_id = user.get("id")
                if not user_id:
                    raise HTTPException(401, "Invalid or expired token")
                return {
                    "sub": user_id,
                    "email": user.get("email"),
                    "aud": "authenticated",
                }
        except Exception:
            # Fall through to local verification below
            pass

    # 2) Local JWT verification fallback (requires configured secret)
    if SUPABASE_JWT_SECRET:
        try:
            from jose import jwt as _jwt

            payload = _jwt.decode(
                token,
                SUPABASE_JWT_SECRET,
                algorithms=["HS256"],
                options={"verify_aud": False},
            )
            if not payload.get("sub"):
                raise ValueError("JWT missing sub claim")
            return payload
        except Exception:
            pass

    raise HTTPException(401, "Invalid or expired token")


def _require_admin(request: Request) -> dict:
    """Validate JWT and verify the caller has the 'admin' role. Raises 403 if not."""
    import json as _json
    import urllib.request

    user = _get_current_user(request)
    user_id = user.get("sub", "")
    if not user_id:
        raise HTTPException(403, "Admin role required")

    supabase_url = SUPABASE_URL
    service_role_key = SUPABASE_SERVICE_ROLE_KEY
    if not supabase_url or not service_role_key:
        raise HTTPException(500, "Server misconfigured: missing Supabase credentials")

    url = (
        f"{supabase_url}/rest/v1/user_roles"
        f"?user_id=eq.{user_id}&role=eq.admin&select=id"
    )
    req = urllib.request.Request(url, headers={
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=5) as resp:
            rows = _json.loads(resp.read())
            if not rows:
                raise HTTPException(403, "Admin role required")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(500, "Failed to verify admin role")

    return user


def _sanitize_filename(filename: str) -> str:
    """Sanitize a user-provided filename to prevent path traversal."""
    basename = os.path.basename(filename)
    return re.sub(r"[^\w.\-]", "_", basename)

# =============================================================================
# SHARED UTILITIES
# =============================================================================

def clean(s):
    return "" if s is None else re.sub(r"\s+", " ", str(s)).strip()

def clean_number(val):
    s = str(val).replace(",", "").strip() if val is not None else ""
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None

def df_to_xml_bytes(df: pd.DataFrame, root_tag="Transactions", row_tag="Transaction") -> bytes:
    root = ET.Element(root_tag)
    for _, row in df.iterrows():
        node = ET.SubElement(root, row_tag)
        for col in df.columns:
            child = ET.SubElement(node, re.sub(r"[^A-Za-z0-9_]", "_", str(col)))
            val = row[col]
            child.text = "" if pd.isna(val) else str(val)
    buf = io.BytesIO()
    ET.ElementTree(root).write(buf, encoding="utf-8", xml_declaration=True)
    return buf.getvalue()

def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Transactions")
        ws = writer.sheets["Transactions"]
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        thin = Side(style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        bold_font = Font(bold=True, size=11)
        normal_font = Font(size=11)
        header_fill = PatternFill(fill_type="solid", fgColor="E8E8E8")

        # Style header row
        for cell in ws[1]:
            cell.font = bold_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Style data rows and compute column widths
        col_widths = [len(str(cell.value or "")) for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for idx, cell in enumerate(row, start=0):
                cell.border = border
                cell.font = normal_font
                cell.alignment = Alignment(vertical="center")
                val_len = len(str(cell.value or ""))
                if val_len > col_widths[idx]:
                    col_widths[idx] = val_len

        # Auto-fit column widths
        for idx, width in enumerate(col_widths, start=1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = min(max(width + 4, 10), 60)

        ws.freeze_panes = "A2"
    return buf.getvalue()

def b64(data: bytes) -> str:
    return base64.b64encode(data).decode()

# =============================================================================
# ── ABAKKUS PMS PARSER ────────────────────────────────────────────────────────
# =============================================================================

DATE_RE_STRICT  = re.compile(r"^\d{2}/\d{2}/\d{4}$")
DATE_RE_ANY     = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
KNOWN_EXCHANGES = {"NSE", "BSE", "MCX", "NCDEX"}
TXN_STARTERS    = ("Buy ", "Sell ", "TDS ")
SUMMARY_TRIGGER = re.compile(r"TRANSACTION\s+STATEMENT\s+SUMMARY", re.IGNORECASE)


def _pms_is_candidate(line: str) -> bool:
    if not any(line.startswith(x) for x in TXN_STARTERS):
        return False
    return DATE_RE_ANY.search(line) is not None


def _pms_parse_tokens(tokens):
    i_date = next((i for i, t in enumerate(tokens) if DATE_RE_STRICT.match(t)), None)
    if i_date is None or i_date + 1 >= len(tokens):
        return None
    tran_date, settlement_date = tokens[i_date], tokens[i_date + 1]
    if not (DATE_RE_STRICT.match(tran_date) and DATE_RE_STRICT.match(settlement_date)):
        return None
    if len(tokens) < i_date + 7:
        return None
    qty, unit_price, brkg, stt, amount = tokens[-5], tokens[-4], tokens[-3], tokens[-2], tokens[-1]
    token_before_qty = tokens[-6] if len(tokens) >= 6 else ""
    if token_before_qty in KNOWN_EXCHANGES:
        exchg = token_before_qty
        security_tokens = tokens[i_date + 2:-6]
    else:
        exchg = ""
        security_tokens = tokens[i_date + 2:-5]
    return {
        "Transaction Description": " ".join(tokens[:i_date]).strip(),
        "Tran Date": tran_date,
        "Settlement Date": settlement_date,
        "Security": " ".join(security_tokens).strip(),
        "Exchg": exchg,
        "Quantity": clean_number(qty),
        "Unit Price": clean_number(unit_price),
        "Brkg.": clean_number(brkg),
        "STT": clean_number(stt),
        "Settlement Amount": clean_number(amount),
    }


def _pms_summary_total(lines):
    total, found = 0.0, False
    for ln in lines:
        ln = ln.strip()
        if not ln or not any(ln.startswith(x) for x in TXN_STARTERS):
            continue
        tokens = ln.split()
        if len(tokens) < 5:
            continue
        v = clean_number(tokens[-1])
        if v is not None:
            total += v
            found = True
    return total if found else None


def parse_pms_pdf(pdf_path: str, password: Optional[str] = None, force_ocr: bool = False):
    rows, problems = [], []
    summary_lines = []
    summary_mode = False

    # Use OCR-fallback text extraction
    pages_text, extraction_method = extract_pdf_pages_text(pdf_path, password, force_ocr=force_ocr)
    total_pages = len(pages_text)

    for page_idx, text in enumerate(pages_text, 1):
        if not text.strip():
            continue
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if any(SUMMARY_TRIGGER.search(l) for l in lines):
            summary_mode = True
        if summary_mode:
            summary_lines.extend(lines)
        i = 0
        while i < len(lines):
            line = lines[i]
            if not _pms_is_candidate(line):
                i += 1
                continue
            combined = line
            row = _pms_parse_tokens(combined.split())
            tries = 0
            while row is None and i + 1 < len(lines) and tries < 2:
                if _pms_is_candidate(lines[i + 1]):
                    break
                combined += " " + lines[i + 1]
                row = _pms_parse_tokens(combined.split())
                i += 1
                tries += 1
            if row:
                rows.append(row)
            else:
                problems.append({"page": page_idx, "line": combined})
            i += 1

    df = pd.DataFrame(rows)
    if not df.empty:
        for col in ["Quantity", "Unit Price", "Brkg.", "STT", "Settlement Amount"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        df["_dt"] = pd.to_datetime(df["Tran Date"], format="%d/%m/%Y", errors="coerce")
        df = df.sort_values(["_dt", "Transaction Description"], kind="stable").drop(columns="_dt")

    summary_total = _pms_summary_total(summary_lines)
    return df, problems, total_pages, summary_total, extraction_method


def add_pms_total_row(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["Transaction Description", "Tran Date", "Settlement Date", "Security",
            "Exchg", "Quantity", "Unit Price", "Brkg.", "STT", "Settlement Amount"]
    if df.empty:
        df = pd.DataFrame(columns=cols)
    else:
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        df = df[cols]
    total_settlement = pd.to_numeric(df["Settlement Amount"], errors="coerce").fillna(0).sum()
    total_row = {c: "" for c in df.columns}
    total_row["Transaction Description"] = "TOTAL"
    total_row["Settlement Amount"] = float(total_settlement)
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

# =============================================================================
# ── KOTAK BANK TEXT-MODE PARSER ───────────────────────────────────────────────
# =============================================================================

from dateutil import parser as dateparser

BANK_DATE_LINE_RE = re.compile(r"^(\d{1,2}\s+[A-Za-z]{3,9},?\s+\d{4})\s+(.*)$")
BANK_MONEY_RE     = re.compile(r"[+-]?\d{1,3}(?:,\d{3})*(?:\.\d{2})")
BANK_REF_RE       = re.compile(
    r"(UPI-\d{10,}|NEFTINW-\d+|IMPS-\d+|ONBF-[A-Za-z0-9]+|\b\d{9,}\b)", re.IGNORECASE
)
BANK_TABLE_HEADER_RE = re.compile(r"date\s+transaction\s+details", re.IGNORECASE)
TABLE_END_RE      = re.compile(
    r"(most\s*important\s*terms|mitc|terms\s*and\s*conditions|schedule\s*of\s*charges"
    r"|grievance|ombudsman|disclaimer|customer\s*care|important\s*information"
    r"|rewards?\s+point|finance\s*charges|interest\s+rate|late\s*payment\s*fee"
    r"|nomination\s+form|registered\s+office|cin\s*:"
    r"|note\s*:\s*this|this\s+is\s+a\s+computer\s+generated"
    r"|e\s*&\s*o\.?\s*e|errors?\s+and\s+omissions)",
    re.IGNORECASE
)
PAGE_FOOTER_RE    = re.compile(r"page\s+\d+\s+of\s+\d+", re.IGNORECASE)


def parse_any_date(s: str):
    s = clean(s)
    if not s:
        return None
    try:
        return dateparser.parse(s, dayfirst=True, fuzzy=True).date().isoformat()
    except Exception:
        return None


def parse_amount(s: str):
    if not s:
        return None
    try:
        return float(clean(s).replace(",", ""))
    except Exception:
        return None


def bank_split_desc_ref(text: str):
    text = clean(text)
    matches = list(BANK_REF_RE.finditer(text))
    if not matches:
        return text, ""
    last = matches[-1]
    return clean(text[:last.start()] + " " + text[last.end():]), clean(last.group(1))


BANK_FOOTER_RE = re.compile(
    r"(kotak|mahindra|bank\s+limited|customer\s+care|download\s+the\s+app"
    r"|registered\s+office|cin\s*:|branch\s+code|ifsc|micr"
    r"|www\.|http|@|toll\s*free|helpline|contact\s+us"
    r"|this\s+is\s+a\s+computer\s+generated|digitally\s+signed"
    r"|note\s*:|important\s*:|disclaimer|subject\s+to"
    r"|e\s*&\s*o\.?\s*e|errors?\s+and\s+omissions"
    r"|nomination\s+|nominee|insurance|mutual\s+fund"
    r"|rbi\b|reserve\s+bank|sebi\b|deposit\s+insurance"
    r"|contents\s+of\s+this|do\s+not\s+share|otp"
    r"|^\*{2,}|^-{3,}|^_{3,}|^\={3,})",
    re.IGNORECASE,
)

def is_bad_bank_line(line: str) -> bool:
    line = clean(line)
    if not line:
        return True
    if PAGE_FOOTER_RE.search(line):
        return True
    if re.search(r"^summary\b", line, re.IGNORECASE):
        return True
    if BANK_FOOTER_RE.search(line):
        return True
    # Lines that are all-caps with no numbers and no date pattern — likely headers/footers
    if line.isupper() and not re.search(r"\d", line) and len(line) > 10:
        return True
    # Lines with long digit-only strings (account numbers, IDs in footers)
    if re.search(r"\b\d{12,}\b", line) and not BANK_DATE_LINE_RE.match(line):
        return True
    return False


def _looks_like_kotak_txn_line(line: str) -> bool:
    if not BANK_DATE_LINE_RE.match(line):
        return False
    return len(BANK_MONEY_RE.findall(line)) >= 2


def parse_kotak_bank(pdf_path: str, password: Optional[str] = None, force_ocr: bool = False) -> pd.DataFrame:
    rows = []
    capture = False

    # Use OCR-fallback text extraction
    pages_text, extraction_method = extract_pdf_pages_text(pdf_path, password, force_ocr=force_ocr)
    total_pages = len(pages_text)

    for txt in pages_text:
        for raw in (txt or "").splitlines():
            line = clean(raw)
            if not line:
                continue
            if not capture and BANK_TABLE_HEADER_RE.search(line):
                capture = True
                continue

            if not capture and _looks_like_kotak_txn_line(line):
                capture = True

            if capture and (re.search(r"^SUMMARY\b", line, re.IGNORECASE) or TABLE_END_RE.search(line)):
                capture = False
                break
            if not capture:
                continue
            m = BANK_DATE_LINE_RE.match(line)
            if not m:
                if rows and not is_bad_bank_line(line):
                    rows[-1]["Description"] = clean(rows[-1]["Description"] + " " + line)
                continue
            dt_raw, rest = m.group(1), m.group(2)
            txn_date = parse_any_date(dt_raw.replace(",", ""))
            monies = BANK_MONEY_RE.findall(rest)
            if len(monies) < 2:
                continue
            txn_amt = parse_amount(monies[-2])
            balance = parse_amount(monies[-1])
            temp = rest
            for tok in (monies[-1], monies[-2]):
                pos = temp.rfind(tok)
                if pos != -1:
                    temp = clean(temp[:pos] + " " + temp[pos + len(tok):])
            desc, ref = bank_split_desc_ref(temp)
            debit = credit = None
            if txn_amt is not None:
                if txn_amt < 0:
                    debit = abs(txn_amt)
                else:
                    credit = txn_amt
            rows.append({"TxnDate": txn_date, "Description": desc, "Reference": ref,
                         "Debit": debit, "Credit": credit, "Balance": balance})
    df = pd.DataFrame(rows)
    if not df.empty:
        for c in ["Debit", "Credit", "Balance"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df, total_pages, extraction_method

# =============================================================================
# ── KOTAK CREDIT CARD OCR PARSER ─────────────────────────────────────────────
# =============================================================================

CC_DATE_RE    = re.compile(r"^(\d{1,2})[\/\-\s](\d{1,2})[\/\-\s](\d{2,4})\s+(.*)$")
AMOUNT_RE     = re.compile(r"\d[\d,]*\.\d{2}")
AMOUNT_ONLY   = re.compile(r"^\s*(\d[\d,]*\.\d{2})\s*(Cr|CR|C\s*r)?\s*$")
FX_ONLY_RE    = re.compile(
    r"^\s*\(?\s*\d[\d,]*\.\d{2}\s*(EUR|USD|GBP|AED|SGD|CAD|AUD|CHF|JPY|CNY)\s*\)?\s*$",
    re.IGNORECASE
)
SECTION_HDR_RE = re.compile(
    r"^(payments?\s+and\s+other\s+credits|primary\s+card\s+transactions.*|"
    r"retail\s+purchases\s+and\s+cash\s+transactions|other\s+charges|"
    r"other\s+fees\s+and\s+charges|fees\s+and\s+charges|finance\s+charges|"
    r"interest\s+charges|total\s+purchase\s*&\s*other\s+charges|"
    r"total\s+purchase\s+&\s+other\s+charges|total\s+due|"
    r"previous\s+balance|new\s+charges|adjustments)$",
    re.IGNORECASE
)
FOOTER_BLOCK_RE = re.compile(
    r"(pay\s+your\s+credit\s+card\s+bills|what\s+you\s+must\s+know|"
    r"kotak\s+mahindra\s+bank\s+limited|contact\s+us|customer\s+care|"
    r"download\s+the\s+app|mobile\s+banking|net\s+banking|"
    r"\*?\s*sms\s+emi|convert\s+your\s+transactions\s+into\s+emi|"
    r"payment\s+of\s+only\s+minimum\s+dues|outstanding\s+balances|"
    r"terms\s+and\s+conditions|disclaimer|grievance|ombudsman|"
    r"registered\s+office|cin\s*:|important\s+information)",
    re.IGNORECASE
)
TXN_HEADER_RE  = re.compile(r"(transaction\s+details|date\s+.*description\s+.*amount)", re.IGNORECASE)
# Additional pattern to detect statement summary/total lines to skip
SUMMARY_LINE_RE = re.compile(
    r"(total\s+amount\s+due|minimum\s+amount\s+due|last\s+payment|"
    r"opening\s+balance|closing\s+balance|credit\s+limit|available\s+limit|"
    r"reward\s+point|statement\s+date|payment\s+due\s+date|"
    r"statement\s+period|billing\s+cycle)",
    re.IGNORECASE
)


def _estimate_gs_timeout(page_count: int) -> int:
    upper_bound = max(CONVERT_TIMEOUT_SECS - 20, GS_TIMEOUT_BASE_SECS)
    if page_count <= 0:
        return GS_TIMEOUT_BASE_SECS
    return min(upper_bound, max(GS_TIMEOUT_BASE_SECS, page_count * 20))


def _normalize_ocr_text(text: str) -> str:
    if not text:
        return ""

    normalized_lines = []
    for raw_line in text.splitlines():
        line = raw_line.replace("₹", " ").replace("|", " ").replace("¦", " ")
        line = line.replace("—", "-").replace("–", "-")
        line = re.sub(r"\s+", " ", line).strip()
        if not line:
            continue

        line = re.sub(r"(?<=\d)[oO](?=\d)", "0", line)
        line = re.sub(r"(?<=\d)[lI](?=\d)", "1", line)
        line = re.sub(r"(?<=\d)[sS](?=\d)", "5", line)
        line = re.sub(r"(\d{1,2})[.\s](\d{1,2})[.\s](\d{2,4})(?=\b)", r"\1/\2/\3", line)
        normalized_lines.append(line)

    return "\n".join(normalized_lines)


def _ocr_quality_score(text: str) -> int:
    if not text:
        return 0
    compact = text.replace("\n", " ").strip()
    char_count = len(compact)
    date_hits = len(OCR_DATE_HINT_RE.findall(text))
    amount_hits = len(OCR_AMOUNT_HINT_RE.findall(text))
    return char_count + (date_hits * 60) + (amount_hits * 20)


def _ocr_image_to_text(img: Image.Image) -> str:
    primary_raw = pytesseract.image_to_string(
        img,
        config="--oem 3 --psm 4 -c preserve_interword_spaces=1",
        timeout=35,
    ) or ""
    primary = _normalize_ocr_text(primary_raw)

    if _ocr_quality_score(primary) >= 140 and OCR_DATE_HINT_RE.search(primary):
        return primary

    fallback_raw = pytesseract.image_to_string(
        img,
        config="--oem 3 --psm 6 -c preserve_interword_spaces=1",
        timeout=35,
    ) or ""
    fallback = _normalize_ocr_text(fallback_raw)

    return fallback if _ocr_quality_score(fallback) > _ocr_quality_score(primary) else primary


def _ocr_pdf_batch(
    pdf_path: str,
    gs_path: str,
    dpi: int,
    gs_timeout: int,
    password: Optional[str] = None,
) -> list[str]:
    """Fallback OCR mode: rasterize all pages in one Ghostscript run."""
    with tempfile.TemporaryDirectory() as tmp:
        out_pattern = os.path.join(tmp, "page_%03d.png")
        cmd = [
            gs_path,
            "-q",
            "-dSAFER",
            "-dBATCH",
            "-dNOPAUSE",
            f"-r{dpi}",
            "-sDEVICE=png16m",
            f"-sOutputFile={out_pattern}",
        ]
        if password:
            cmd.append(f"-sPDFPassword={password}")
        cmd.append(pdf_path)

        subprocess.run(
            cmd,
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=gs_timeout,
        )

        image_files = sorted(f for f in os.listdir(tmp) if f.endswith(".png"))
        if not image_files:
            raise RuntimeError("Ghostscript produced no rasterized pages for OCR")

        pages_text = []
        for fn in image_files:
            image_path = os.path.join(tmp, fn)
            try:
                with Image.open(image_path) as img:
                    text = _ocr_image_to_text(img)
                pages_text.append(text)
            except RuntimeError:
                pages_text.append("")

        return pages_text


PDF_OCR_CHUNK_SIZE = 10  # pages per chunk for OCR splitting


def _split_pdf_for_ocr(pdf_path: str, chunk_size: int, password: Optional[str] = None) -> list[str]:
    """Split a large PDF into temp chunk files for OCR. Returns list of paths (cleaned up by caller)."""
    if not PIKEPDF_AVAILABLE:
        return [pdf_path]
    try:
        open_kw = {"password": password} if password else {}
        src = pikepdf.open(pdf_path, **open_kw)
    except Exception:
        return [pdf_path]

    total = len(src.pages)
    if total <= chunk_size:
        src.close()
        return [pdf_path]

    chunk_paths = []
    for start in range(0, total, chunk_size):
        end = min(start + chunk_size, total)
        chunk_pdf = pikepdf.Pdf.new()
        for i in range(start, end):
            chunk_pdf.pages.append(src.pages[i])
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        chunk_pdf.save(tmp.name)
        chunk_pdf.close()
        chunk_paths.append(tmp.name)
    src.close()
    return chunk_paths


def ocr_pdf(pdf_path: str, password: Optional[str] = None) -> list[str]:
    """OCR all pages of a PDF, splitting into chunks for large files."""
    if not OCR_AVAILABLE:
        raise RuntimeError("pytesseract / Pillow not installed")

    gs_path = shutil.which(GS_EXE) or GS_EXE
    tess_path = shutil.which(TESSERACT_EXE) or TESSERACT_EXE
    if not shutil.which(gs_path) and not os.path.exists(gs_path):
        raise RuntimeError(f"Ghostscript not found: {GS_EXE}")
    if not shutil.which(tess_path) and not os.path.exists(tess_path):
        raise RuntimeError(f"Tesseract not found: {TESSERACT_EXE}")
    pytesseract.pytesseract.tesseract_cmd = tess_path

    page_count = 0
    try:
        open_kwargs = {"path_or_fp": pdf_path}
        if password:
            open_kwargs["password"] = password
        with pdfplumber.open(**open_kwargs) as pdf:
            page_count = len(pdf.pages)
    except Exception:
        page_count = 0

    dpi = OCR_DPI_LARGE_DOC if page_count >= OCR_HIGH_PAGE_COUNT else OCR_DPI

    # Split large PDFs into chunks for reliable OCR
    chunk_paths = _split_pdf_for_ocr(pdf_path, PDF_OCR_CHUNK_SIZE, password)
    all_pages_text = []

    for chunk_path in chunk_paths:
        try:
            chunk_page_count = 0
            try:
                with pdfplumber.open(chunk_path) as cpdf:
                    chunk_page_count = len(cpdf.pages)
            except Exception:
                chunk_page_count = PDF_OCR_CHUNK_SIZE
            gs_timeout = _estimate_gs_timeout(chunk_page_count)
            # Don't pass password for split chunks (already decrypted by pikepdf)
            chunk_pw = password if chunk_path == pdf_path else None
            pages = _ocr_pdf_batch(chunk_path, gs_path, dpi, gs_timeout, chunk_pw)
            all_pages_text.extend(pages)
        finally:
            if chunk_path != pdf_path:
                try:
                    os.unlink(chunk_path)
                except Exception:
                    pass

    return all_pages_text


def _safe_rows_json(df: pd.DataFrame) -> list[dict]:
    """Serialize dataframe rows safely for JSON (NaN/Inf -> None)."""
    import math

    rows_json = df.where(df.notna(), other=None).to_dict(orient="records")
    for row in rows_json:
        for k, v in row.items():
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                row[k] = None
    return rows_json




def extract_pdf_pages_text(pdf_path: str, password: Optional[str] = None,
                           x_tolerance: int = 2, y_tolerance: int = 2,
                           force_ocr: bool = False) -> tuple[list[str], str]:
    """
    Extract text from each page of a PDF using pdfplumber.
    Falls back to OCR if pdfplumber yields very little usable text.
    Returns (list_of_page_texts, method) where method is 'text' or 'ocr'.
    """
    open_kwargs = {"path_or_fp": pdf_path}
    if password:
        open_kwargs["password"] = password

    pages_text = []
    try:
        with pdfplumber.open(**open_kwargs) as pdf:
            for page in pdf.pages:
                text = page.extract_text(x_tolerance=x_tolerance, y_tolerance=y_tolerance) or ""
                pages_text.append(text)
    except Exception:
        pages_text = []

    # Check if pdfplumber yielded enough text — if not, fall back to OCR
    total_chars = sum(len(t.strip()) for t in pages_text)
    has_date = any(re.search(r"\d{2}[\/\-]\d{2}[\/\-]\d{2,4}", t) for t in pages_text if t)
    should_try_ocr = force_ocr or total_chars < 100 or not has_date
    if should_try_ocr:
        if OCR_AVAILABLE and (shutil.which(GS_EXE) or os.path.exists(GS_EXE)):
            try:
                ocr_pages = ocr_pdf(pdf_path, password)
                ocr_pages = [_normalize_ocr_text(t) for t in ocr_pages]
                ocr_chars = sum(len(t.strip()) for t in ocr_pages)
                ocr_has_date = any(re.search(r"\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}", t) for t in ocr_pages if t)

                if force_ocr and ocr_pages:
                    return ocr_pages, "ocr"

                if ocr_pages and ocr_chars >= max(int(total_chars * 0.8), 60) and (ocr_has_date or not has_date):
                    return ocr_pages, "ocr"
            except Exception:
                pass

    return pages_text, "text"


def extract_cc_blocks(pages_text: list[str]) -> list[list[str]]:
    tables, current, in_table = [], [], False
    for txt in pages_text:
        for raw in (txt or "").splitlines():
            line = clean(raw)
            if not line:
                continue
            if not in_table and TXN_HEADER_RE.search(line):
                in_table, current = True, []
                continue
            if in_table:
                if TABLE_END_RE.search(line) or FOOTER_BLOCK_RE.search(line):
                    if current:
                        tables.append(current)
                    in_table, current = False, []
                    continue
                if PAGE_FOOTER_RE.search(line):
                    continue
                current.append(line)
    if in_table and current:
        tables.append(current)
    return tables


def parse_cc_lines(lines: list[str]) -> pd.DataFrame:
    rows, last_date, pending_section, pending_label = [], None, "", ""
    for raw in lines:
        line = clean(raw)
        if not line or FOOTER_BLOCK_RE.search(line) or PAGE_FOOTER_RE.search(line):
            continue
        # Skip summary/info lines
        if SUMMARY_LINE_RE.search(line):
            continue
        if SECTION_HDR_RE.match(line):
            pending_section = line
            rows.append({"TxnDate": last_date, "TransactionDetails": line,
                         "SpendsArea": "", "Debit": None, "Credit": None})
            continue
        if FX_ONLY_RE.match(line):
            if rows:
                rows[-1]["TransactionDetails"] = clean(rows[-1]["TransactionDetails"] + " " + line)
            continue
        m = CC_DATE_RE.match(line)
        if m:
            dd, mm, yyyy, rest = m.group(1), m.group(2), m.group(3), m.group(4)
            # Handle 2-digit years
            if len(yyyy) == 2:
                yyyy = "20" + yyyy if int(yyyy) < 50 else "19" + yyyy
            txn_date = parse_any_date(f"{dd}/{mm}/{yyyy}")
            if txn_date is None:
                # Try swapping dd/mm in case OCR misread
                txn_date = parse_any_date(f"{mm}/{dd}/{yyyy}")
            last_date, pending_label = txn_date, ""
            amts = AMOUNT_RE.findall(rest)
            if not amts:
                pending_label = clean(rest)
                continue
            amt = parse_amount(amts[-1])
            if amt is None:
                continue
            is_cr = bool(re.search(r"\bCr\b|\bCR\b|\bcredit\b", rest, re.IGNORECASE))
            rest_c = re.sub(r"\bCr\b|\bCR\b", "", rest).strip()
            pos = rest_c.rfind(amts[-1])
            left = clean(rest_c[:pos]) if pos != -1 else clean(rest_c)
            toks = left.split()
            if toks and toks[-1].isalpha() and len(toks[-1]) <= 20:
                spends_area, details = toks[-1], clean(" ".join(toks[:-1]))
            else:
                spends_area, details = "", left
            # Don't prepend section header to details — keep them clean
            rows.append({"TxnDate": txn_date, "TransactionDetails": details,
                         "SpendsArea": spends_area,
                         "Debit": None if is_cr else amt,
                         "Credit": amt if is_cr else None})
            continue
        mo = AMOUNT_ONLY.match(line)
        if mo and last_date:
            amt = parse_amount(mo.group(1))
            if amt is None or not pending_label:
                if rows:
                    rows[-1]["TransactionDetails"] = clean(rows[-1]["TransactionDetails"] + " " + line)
                continue
            is_cr = bool(mo.group(2))
            label = clean(pending_label)
            pending_label = ""
            rows.append({"TxnDate": last_date, "TransactionDetails": label, "SpendsArea": "",
                         "Debit": None if is_cr else amt, "Credit": amt if is_cr else None})
            continue
        amts2 = AMOUNT_RE.findall(line)
        if amts2 and last_date:
            if re.search(r"\b(EUR|USD|GBP|AED|SGD|CAD)\b", line, re.IGNORECASE):
                if rows:
                    rows[-1]["TransactionDetails"] = clean(rows[-1]["TransactionDetails"] + " " + line)
                continue
            amt = parse_amount(amts2[-1])
            if amt is None:
                continue
            is_cr = bool(re.search(r"\bCr\b|\bCR\b|\bcredit\b", line, re.IGNORECASE))
            lc = re.sub(r"\bCr\b|\bCR\b", "", line).strip()
            pos = lc.rfind(amts2[-1])
            left = clean(lc[:pos]) if pos != -1 else clean(lc)
            rows.append({"TxnDate": last_date, "TransactionDetails": left or line, "SpendsArea": "",
                         "Debit": None if is_cr else amt, "Credit": amt if is_cr else None})
            continue
        # Only merge short text into pending_label — avoid absorbing garbage
        if len(line.split()) <= 6 and re.search(r"[A-Za-z]", line) and not re.search(r"\d{5,}", line):
            pending_label = clean((pending_label + " " + line).strip())
            continue
        if rows:
            rows[-1]["TransactionDetails"] = clean(rows[-1]["TransactionDetails"] + " " + line)

    df = pd.DataFrame(rows)
    if not df.empty:
        for c in ["Debit", "Credit"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def format_cc_output(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["Date", "Transaction Details", "Spends Area", "Debit", "Credit"]
    if df.empty:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame()
    out["Date"]                = df.get("TxnDate", "").astype(str).apply(clean)
    out["Transaction Details"] = df.get("TransactionDetails", "").astype(str).apply(clean)
    out["Spends Area"]         = df.get("SpendsArea", "").astype(str).apply(clean)
    out["Debit"]               = pd.to_numeric(df.get("Debit"), errors="coerce")
    out["Credit"]              = pd.to_numeric(df.get("Credit"), errors="coerce")
    return out[cols]

# =============================================================================
# ── HDFC BANK STATEMENT PARSER ───────────────────────────────────────────────
# =============================================================================

# =============================================================================
# ── HDFC BANK STATEMENT PARSER (with built-in OCR fallback) ──────────────────
# =============================================================================

HDFC_DATE_START_RE = re.compile(r"^\d{2}/\d{2}/\d{2}")
HDFC_DATE_ANY_RE   = re.compile(r"\d{2}/\d{2}/\d{2}")
HDFC_AMOUNT_RE     = re.compile(r"(\d{1,3}(?:,\d{3})*\.\d{2})")

HDFC_HEADER_TEXT       = "Date Narration Chq./Ref.No. ValueDt WithdrawalAmt. DepositAmt. ClosingBalance"
HDFC_FOOTER_START      = "HDFCBANKLIMITED"
HDFC_SUMMARY_START     = "STATEMENTSUMMARY:-"       # Digital (no spaces)
HDFC_SUMMARY_START_OCR = "STATEMENT SUMMARY :-"     # OCR (with spaces)


def _hdfc_is_standalone_ref(token: str, is_ocr: bool) -> bool:
    if "-" in token:
        return False
    if is_ocr:
        return bool(re.fullmatch(r"[A-Za-z0-9.,]{15,}", token))
    return bool(re.fullmatch(r"[A-Za-z0-9]{15,}", token))


def _hdfc_clean_amount(x):
    if isinstance(x, str):
        return float(x.replace(",", ""))
    return float(x or 0.0)


def _hdfc_extract_text_hybrid(pdf_path: str, password: Optional[str] = None, force_ocr: bool = False):
    """Try pdfplumber first, fall back to OCR only if text is insufficient."""
    text = ""
    is_ocr = False
    total_pages = 0

    try:
        open_kwargs = {"path_or_fp": pdf_path}
        if password:
            open_kwargs["password"] = password
        with pdfplumber.open(**open_kwargs) as pdf:
            total_pages = len(pdf.pages)
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n"
    except Exception:
        pass

    # Only fall back to OCR if pdfplumber yields very little text (or when forced)
    if force_ocr or not text or len(text.strip()) < 100 or "Date" not in text:
        is_ocr = True
        if OCR_AVAILABLE and (shutil.which(GS_EXE) or os.path.exists(GS_EXE)) and (shutil.which(TESSERACT_EXE) or os.path.exists(TESSERACT_EXE)):
            try:
                pages_text = ocr_pdf(pdf_path, password)
                total_pages = len(pages_text)
                text = "\n".join(_normalize_ocr_text(p) for p in pages_text)
            except Exception:
                pass

    method = "ocr" if is_ocr else "text"
    return text, is_ocr, total_pages, method


def _hdfc_extract_blocks(text: str, is_ocr: bool):
    """Extract transaction blocks with dual hard-stop logic for digital vs OCR."""
    all_blocks = []
    current_block = ""
    start_recording = False
    in_footer_wait = False

    for line in text.splitlines():
        clean_line = line.strip()
        if not clean_line:
            continue

        # 1. Recording start logic — for digital, skip repeated header lines
        if not is_ocr:
            if HDFC_HEADER_TEXT.replace(" ", "") in clean_line.replace(" ", ""):
                start_recording = True
                continue
        elif not start_recording:
            if HDFC_DATE_START_RE.match(clean_line):
                start_recording = True

        # 2. Dual hard-stop logic
        if is_ocr:
            if HDFC_SUMMARY_START_OCR in clean_line:
                if current_block:
                    all_blocks.append(current_block)
                break
        else:
            if HDFC_SUMMARY_START in clean_line.replace(" ", ""):
                if current_block:
                    all_blocks.append(current_block)
                break

        # 3. Footer skip logic
        if HDFC_FOOTER_START in clean_line.replace(" ", ""):
            in_footer_wait = True
            continue

        if in_footer_wait:
            if HDFC_DATE_START_RE.match(clean_line):
                in_footer_wait = False
            else:
                continue

        # 4. Block accumulation
        if start_recording:
            if HDFC_DATE_START_RE.match(clean_line):
                if current_block:
                    all_blocks.append(current_block)
                current_block = clean_line
            elif current_block:
                current_block += " " + clean_line

    # Flush last block if loop ended without hitting summary marker (big statements)
    if current_block:
        all_blocks.append(current_block)

    return all_blocks


def _hdfc_parse_row(block_text: str, prev_balance, is_ocr: bool):
    tokens = block_text.split()
    if not tokens:
        return None, prev_balance

    txn_date = tokens[0]
    ref_no = next((t for t in tokens if _hdfc_is_standalone_ref(t, is_ocr)), "")

    amounts = HDFC_AMOUNT_RE.findall(block_text)
    withdrawal = deposit = closing_bal = 0.0
    needs_manual_check = False

    if len(amounts) >= 2:
        closing_bal = _hdfc_clean_amount(amounts[-1])
        txn_val = _hdfc_clean_amount(amounts[-2])

        if "SELF - CHQ PAID" in block_text.upper() or " DR" in block_text.upper():
            withdrawal = txn_val
        elif " CR" in block_text.upper():
            deposit = txn_val
        elif prev_balance:
            if closing_bal > prev_balance:
                deposit = txn_val
            else:
                withdrawal = txn_val
        else:
            needs_manual_check = True

    # Build narration by removing dates, amounts, ref
    narration = block_text
    for d in HDFC_DATE_ANY_RE.findall(block_text):
        narration = narration.replace(d, "", 1)
    for a in amounts:
        narration = narration.replace(a, "", 1)
    if ref_no:
        narration = narration.replace(ref_no, "", 1)
    narration = re.sub(r"\s+", " ", narration).strip()

    return {
        "Date": txn_date,
        "Narration": narration,
        "Chq./Ref.No": ref_no,
        "WithdrawalAmt": withdrawal,
        "DepositAmt": deposit,
        "ClosingBalance": closing_bal,
        "Flag": "MANUAL_VERIFY" if needs_manual_check else "",
    }, closing_bal


HDFC_CHUNK_SIZE = 10  # pages per chunk for big PDFs


def _split_pdf_into_chunks(pdf_path: str, chunk_size: int, password: Optional[str] = None) -> list:
    """Split a PDF into smaller temp files of chunk_size pages each. Returns list of temp file paths."""
    if not PIKEPDF_AVAILABLE:
        return [pdf_path]

    try:
        open_kwargs = {}
        if password:
            open_kwargs["password"] = password
        src = pikepdf.open(pdf_path, **open_kwargs)
    except Exception:
        return [pdf_path]

    total = len(src.pages)
    if total <= chunk_size:
        src.close()
        return [pdf_path]

    chunk_paths = []
    for start in range(0, total, chunk_size):
        end = min(start + chunk_size, total)
        chunk_pdf = pikepdf.Pdf.new()
        for i in range(start, end):
            chunk_pdf.pages.append(src.pages[i])
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        chunk_pdf.save(tmp.name)
        chunk_pdf.close()
        chunk_paths.append(tmp.name)

    src.close()
    return chunk_paths


def parse_hdfc_bank(pdf_path: str, password: Optional[str] = None, force_ocr: bool = False):
    # Count total pages first
    total_pages = 0
    try:
        open_kwargs = {"path_or_fp": pdf_path}
        if password:
            open_kwargs["password"] = password
        with pdfplumber.open(**open_kwargs) as pdf:
            total_pages = len(pdf.pages)
    except Exception:
        pass

    # For big PDFs, split into chunks, parse each, merge results
    if total_pages > HDFC_CHUNK_SIZE and PIKEPDF_AVAILABLE:
        chunk_paths = _split_pdf_into_chunks(pdf_path, HDFC_CHUNK_SIZE, password)
    else:
        chunk_paths = [pdf_path]

    all_rows = []
    all_problem_lines = []
    running_balance = 0.0
    combined_method = "text"
    combined_metadata = {"opening_balance": 0.0, "exp_dr_sum": 0.0, "exp_cr_sum": 0.0}

    for idx, chunk_path in enumerate(chunk_paths):
        try:
            full_text, is_ocr, _cp, method = _hdfc_extract_text_hybrid(
                chunk_path, password if chunk_path == pdf_path else None, force_ocr=force_ocr
            )
            if is_ocr:
                combined_method = "ocr"

            # Only extract opening balance from the first chunk
            if idx == 0:
                open_bal_match = re.search(r"Opening\s+Balance\s+([\d,]+\.\d{2})", full_text, re.IGNORECASE)
                running_balance = _hdfc_clean_amount(open_bal_match.group(1)) if open_bal_match else 0.0
                combined_metadata["opening_balance"] = running_balance

                # Extract summary metadata from first chunk text
                summary_pattern = re.compile(r"(\d+)\s+(\d+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})")
                summary_matches = summary_pattern.findall(full_text)
                if summary_matches:
                    last_match = summary_matches[-1]
                    combined_metadata["exp_dr_sum"] = _hdfc_clean_amount(last_match[2])
                    combined_metadata["exp_cr_sum"] = _hdfc_clean_amount(last_match[3])

            blocks = _hdfc_extract_blocks(full_text, is_ocr)

            for b in blocks:
                row, running_balance = _hdfc_parse_row(b, running_balance, is_ocr)
                if row:
                    all_rows.append(row)
                else:
                    all_problem_lines.append({"page": 0, "line": b[:120]})

            # Also check later chunks for summary metadata
            if idx > 0:
                summary_pattern = re.compile(r"(\d+)\s+(\d+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})")
                summary_matches = summary_pattern.findall(full_text)
                if summary_matches:
                    last_match = summary_matches[-1]
                    combined_metadata["exp_dr_sum"] = _hdfc_clean_amount(last_match[2])
                    combined_metadata["exp_cr_sum"] = _hdfc_clean_amount(last_match[3])
        finally:
            # Clean up temp chunk files
            if chunk_path != pdf_path:
                try:
                    os.unlink(chunk_path)
                except Exception:
                    pass

    df = pd.DataFrame(all_rows)
    return df, all_problem_lines, total_pages, combined_metadata, combined_method


def add_hdfc_total_row(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["Date", "Narration", "Chq./Ref.No", "WithdrawalAmt", "DepositAmt", "ClosingBalance", "Flag"]
    if df.empty:
        df = pd.DataFrame(columns=cols)
    else:
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        df = df[cols]
    wd = pd.to_numeric(df["WithdrawalAmt"], errors="coerce").fillna(0).sum()
    dp = pd.to_numeric(df["DepositAmt"], errors="coerce").fillna(0).sum()
    total_row = {c: "" for c in cols}
    total_row["Narration"] = "TOTAL"
    total_row["WithdrawalAmt"] = float(wd)
    total_row["DepositAmt"] = float(dp)
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

# =============================================================================
# ── ICICI BANK WEALTH MANAGEMENT PARSER ──────────────────────────────────────
# =============================================================================

ICICI_DATE_RE          = re.compile(r"^\d{2}-\d{2}-\d{4}")
ICICI_DATE_ANYWHERE_RE = re.compile(r"\d{2}-\d{2}-\d{4}")
ICICI_AMOUNT_RE        = re.compile(r"(\d[\d,]*\.\d{2})")
ICICI_PAGE_FOOTER_RE   = re.compile(r"Page\s+\d+\s+of\s+\d+", re.IGNORECASE)
ICICI_HEADER_PATTERN   = re.compile(r"DATE\s+MODE.*PARTICULARS", re.IGNORECASE)
ICICI_TOTAL_STOP       = "Total:"
ICICI_PRIMARY_FOOTER   = "This is an authenticated intimation"
ICICI_FINAL_STOP       = "Account Related Other Information"


def _icici_clean_amount(x):
    if isinstance(x, str):
        x = x.replace(",", "")
        try:
            return float(x)
        except Exception:
            return 0.0
    return float(x or 0.0)


def _icici_extract_opening_balance(text: str):
    match = re.search(r"B/F\s+([\d,]+\.\d{2})", text)
    if match:
        return _icici_clean_amount(match.group(1))
    return 0.0


def _icici_extract_raw_blocks(text: str):
    all_rows = []
    current_block = ""
    continuation_lines = []
    start_recording = False
    pending_particulars = None

    for line in text.splitlines():
        clean_line = line.strip()
        if not clean_line:
            continue

        # STOP CONDITIONS
        if ICICI_FINAL_STOP in clean_line or ICICI_PRIMARY_FOOTER in clean_line or ICICI_PAGE_FOOTER_RE.search(clean_line):
            start_recording = False
            if ICICI_FINAL_STOP in clean_line:
                if current_block:
                    all_rows.append((current_block, continuation_lines))
                return all_rows
            continue

        # START
        if ICICI_HEADER_PATTERN.search(clean_line):
            start_recording = True
            continue

        if ICICI_TOTAL_STOP in clean_line:
            start_recording = False
            continue

        if not start_recording:
            continue

        # NEFT / RTGS
        first_word = clean_line.upper().split()[0]
        if first_word.startswith(("NEFT", "RTGS")):
            pending_particulars = clean_line
            continue

        # DATE START
        if ICICI_DATE_RE.match(clean_line):
            if current_block:
                all_rows.append((current_block, continuation_lines))
            continuation_lines = []
            if pending_particulars:
                current_block = pending_particulars + " " + clean_line
                pending_particulars = None
            else:
                current_block = clean_line
            continue

        # CONTINUATION
        if current_block:
            continuation_lines.append(clean_line)

    if current_block:
        all_rows.append((current_block, continuation_lines))

    return all_rows


def _icici_parse_to_dataframe(raw_rows, initial_bal):
    data = []
    current_bal = initial_bal

    for row, continuation_lines in raw_rows:
        pending_particulars = ""

        # PREFIX (NEFT etc.)
        date_match_full = ICICI_DATE_ANYWHERE_RE.search(row)
        if not date_match_full:
            continue

        date_index = date_match_full.start()
        prefix = row[:date_index].strip()
        if prefix.upper().startswith(("NEFT", "RTGS")):
            pending_particulars = prefix
            row = row[date_index:]

        # DATE
        date_match = ICICI_DATE_RE.match(row)
        if not date_match:
            continue

        date_val = date_match.group()

        # AMOUNTS
        amounts_found = ICICI_AMOUNT_RE.findall(row)

        # B/F ROW
        if "B/F" in row and len(amounts_found) >= 1:
            current_bal = _icici_clean_amount(amounts_found[-1])
            data.append({
                "DATE": date_val,
                "MODE": "",
                "PARTICULARS": "B/F",
                "DEPOSITS": 0.0,
                "WITHDRAWALS": 0.0,
                "BALANCE": current_bal
            })
            continue

        # SAFETY
        if len(amounts_found) < 2:
            continue

        new_bal = _icici_clean_amount(amounts_found[-1])
        txn_amt = _icici_clean_amount(amounts_found[-2])

        # TOKENIZE
        remaining = row[len(date_val):].strip()
        tokens = remaining.split()

        # MODE
        mode_val = ""
        particulars_start_idx = 0
        if len(tokens) >= 2:
            w1, w2 = tokens[0], tokens[1]
            if (w1.isalpha() or w1.isdigit()) and (w2.isalpha() or w2.isdigit()):
                mode_val = f"{w1} {w2}"
                particulars_start_idx = 2

        # MAIN PARTICULARS
        particulars_list = []
        for t in tokens[particulars_start_idx:]:
            if t in amounts_found:
                break
            particulars_list.append(t)

        main_particulars = " ".join(particulars_list).strip()

        # FINAL ORDER BUILD
        final_particulars = main_particulars

        # 1. NEFT FIRST
        if pending_particulars:
            final_particulars = pending_particulars + " " + final_particulars

        # 2. CONTINUATION LAST
        if continuation_lines:
            continuation_text = " ".join(continuation_lines)
            final_particulars = final_particulars + " " + continuation_text

        # CREDIT/DEBIT
        deposit, withdrawal = 0.0, 0.0
        if new_bal > current_bal:
            deposit = txn_amt
        else:
            withdrawal = txn_amt

        data.append({
            "DATE": date_val,
            "MODE": mode_val,
            "PARTICULARS": final_particulars.strip(),
            "DEPOSITS": deposit,
            "WITHDRAWALS": withdrawal,
            "BALANCE": new_bal
        })

        current_bal = new_bal

    return pd.DataFrame(data)


def parse_icici_bank(pdf_path: str, password: Optional[str] = None, force_ocr: bool = False):
    pages_text, extraction_method = extract_pdf_pages_text(pdf_path, password, force_ocr=force_ocr)
    total_pages = len(pages_text)
    full_text = "\n".join(pages_text)

    opening_balance = _icici_extract_opening_balance(full_text)
    raw_rows = _icici_extract_raw_blocks(full_text)
    df = _icici_parse_to_dataframe(raw_rows, opening_balance)

    problems = []
    return df, problems, total_pages, extraction_method


def add_icici_total_row(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["DATE", "MODE", "PARTICULARS", "DEPOSITS", "WITHDRAWALS", "BALANCE"]
    if df.empty:
        df = pd.DataFrame(columns=cols)
    else:
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        df = df[cols]
    dep = pd.to_numeric(df["DEPOSITS"], errors="coerce").fillna(0).sum()
    wd = pd.to_numeric(df["WITHDRAWALS"], errors="coerce").fillna(0).sum()
    total_row = {c: "" for c in cols}
    total_row["PARTICULARS"] = "TOTAL"
    total_row["DEPOSITS"] = float(dep)
    total_row["WITHDRAWALS"] = float(wd)
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


# =============================================================================
# ── AXIS CREDIT CARD PARSER ──────────────────────────────────────────────────
# =============================================================================

AXIS_DATE_RE = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
AXIS_TXN_PATTERN = re.compile(
    r"(\d{2}/\d{2}/\d{4})\s+(.*?)\s+([A-Z &]+)\s+([\d,]+\.\d{2})\s+(Dr|Cr)"
)


def parse_axis_cc(pdf_path: str, password: Optional[str] = None, force_ocr: bool = False):
    rows = []
    problem_lines = []

    # Use OCR-fallback text extraction
    pages_text, extraction_method = extract_pdf_pages_text(pdf_path, password, force_ocr=force_ocr)
    total_pages = len(pages_text)

    for page_idx, text in enumerate(pages_text, 1):
        for line in (text or "").split("\n"):
            match = AXIS_TXN_PATTERN.search(line)
            if match:
                date, desc, category, amount, drcr = match.groups()
                amount_val = float(amount.replace(",", ""))
                if drcr == "Cr":
                    amount_val = -amount_val
                rows.append({
                    "Date": date,
                    "Description": desc.strip(),
                    "Category": category.strip(),
                    "Amount": amount_val,
                    "Type": drcr,
                })
            else:
                if AXIS_DATE_RE.search(line):
                    problem_lines.append({"page": page_idx, "line": line.strip()})
    df = pd.DataFrame(rows)
    if not df.empty:
        df["_dt"] = pd.to_datetime(df["Date"], format="%d/%m/%Y", errors="coerce")
        df = df.sort_values("_dt").drop(columns="_dt")
    return df, problem_lines, total_pages, extraction_method


def add_axis_total_row(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["Date", "Description", "Category", "Amount", "Type"]
    if df.empty:
        df = pd.DataFrame(columns=cols)
    else:
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        df = df[cols]
    total_amount = pd.to_numeric(df["Amount"], errors="coerce").fillna(0).sum()
    total_row = {c: "" for c in cols}
    total_row["Description"] = "TOTAL"
    total_row["Amount"] = float(total_amount)
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


# =============================================================================
# ── /convert ENDPOINT ─────────────────────────────────────────────────────────
# =============================================================================


@app.post("/convert")
async def convert(
    request: Request,
    file: UploadFile = File(...),
    mode: str = Form(...),          # "pms" | "kotak" | "aif" | "axis" | "hdfc" | "icici"
    sub_mode: str = Form(""),       # "bank" | "cc" (only for kotak)
    password: Optional[str] = Form(None),
):
    _get_current_user(request)  # require authentication
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are accepted.")

    # Save upload to input folder
    content = await file.read()

    # ── File size info (no rejection, just log) ────────────────────────
    size_mb = len(content) / (1024 * 1024)

    safe_name = _sanitize_filename(file.filename)
    input_path = INPUT_DIR / safe_name
    input_path.write_bytes(content)
    tmp_path = str(input_path)

    # ── Run the actual processing in a thread with timeout ───────────────
    def _do_convert():
        return _convert_sync(tmp_path, file.filename, mode, sub_mode, password)

    try:
        result = await asyncio.wait_for(
            asyncio.get_event_loop().run_in_executor(None, _do_convert),
            timeout=CONVERT_TIMEOUT_SECS,
        )
        return JSONResponse(result)

    except asyncio.TimeoutError:
        # Force garbage collection to free memory from partial processing
        gc.collect()
        raise HTTPException(
            504,
            f"Processing timed out after {CONVERT_TIMEOUT_SECS} seconds. "
            "The PDF may be too large or complex. Try a file with fewer pages."
        )
    except HTTPException:
        raise
    except MemoryError:
        gc.collect()
        raise HTTPException(
            507,
            "Server ran out of memory processing this PDF. "
            "Try a smaller file or split the PDF into parts."
        )
    except Exception as e:
        gc.collect()
        raise HTTPException(500, f"Extraction failed: {repr(e)}")
    finally:
        try:
            if input_path.exists():
                input_path.unlink()
        except Exception:
            pass


def _convert_sync(tmp_path: str, filename: str, mode: str, sub_mode: str, password: Optional[str]):
    """Synchronous conversion logic — runs in a thread pool."""
    # ── Early check: is the PDF password-protected? ──────────────────────
    try:
        import pikepdf
        try:
            test_pdf = pikepdf.open(tmp_path, password=password or "")
            test_pdf.close()
        except pikepdf._core.PasswordError:
            if not password:
                raise HTTPException(
                    400,
                    "This PDF is password-protected. Please provide the password and try again."
                )
            else:
                raise HTTPException(
                    400,
                    "The password you entered is incorrect. Please check and try again."
                )
    except HTTPException:
        raise
    except ImportError:
        # pikepdf not available, fall back to pdfplumber check
        try:
            with pdfplumber.open(tmp_path) as _test:
                pass
        except Exception as e:
            err_str = str(e).lower()
            if "password" in err_str or "encrypt" in err_str:
                if not password:
                    raise HTTPException(
                        400,
                        "This PDF is password-protected. Please provide the password and try again."
                    )
                else:
                    raise HTTPException(
                        400,
                        "The password you entered is incorrect. Please check and try again."
                    )
    except Exception:
        pass  # Not a password issue, continue normally

    try:
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        problems = []
        df_out   = pd.DataFrame()
        total_pages = 0
        summary_total = None
        detail_total  = 0.0
        status        = "UNKNOWN"
        match_text    = "Summary total not found (cannot compare)."
        method        = "unknown"

        # ── PMS ──────────────────────────────────────────────────────────────
        if mode == "pms":
            df, problems, total_pages, summary_total, ext_method = parse_pms_pdf(tmp_path, password)
            if ext_method == "text" and len(df) < MIN_ROWS:
                df_retry, problems_retry, pages_retry, summary_retry, method_retry = parse_pms_pdf(
                    tmp_path,
                    password,
                    force_ocr=True,
                )
                if len(df_retry) > len(df):
                    df, problems, total_pages, summary_total, ext_method = (
                        df_retry,
                        problems_retry,
                        pages_retry,
                        summary_retry,
                        method_retry,
                    )
            df_with_total = add_pms_total_row(df)
            detail_total  = float(
                pd.to_numeric(df.get("Settlement Amount", pd.Series(dtype=float)), errors="coerce")
                .fillna(0).sum()
            )
            method = f"pms_{ext_method}"
            df_out = df_with_total

            if summary_total is not None:
                diff = detail_total - float(summary_total)
                if abs(diff) <= 0.01:
                    status, match_text = "MATCH", \
                        "Detail TOTAL matches Transaction Summary last-column total (within rounding tolerance)."
                else:
                    status, match_text = "MISMATCH", \
                        "Detail TOTAL does NOT match Transaction Summary last-column total."

        # ── KOTAK BANK ───────────────────────────────────────────────────────
        elif mode == "kotak" and sub_mode == "bank":
            df_bank, total_pages, ext_method = parse_kotak_bank(tmp_path, password)
            if ext_method == "text" and len(df_bank) < MIN_ROWS:
                df_retry, pages_retry, method_retry = parse_kotak_bank(tmp_path, password, force_ocr=True)
                if len(df_retry) > len(df_bank):
                    df_bank, total_pages, ext_method = df_retry, pages_retry, method_retry
            df_out  = df_bank
            method  = f"bank_{ext_method}"
            detail_total = float(
                pd.to_numeric(df_out.get("Credit", pd.Series(dtype=float)), errors="coerce")
                .fillna(0).sum()
            )

        # ── KOTAK CREDIT CARD ────────────────────────────────────────────────
        elif mode == "kotak" and sub_mode == "cc":
            # Kotak CC always uses OCR (scanned statements)
            method = "cc_ocr"
            if not OCR_AVAILABLE:
                raise HTTPException(500, "pytesseract / Pillow not installed on the server.")
            if not shutil.which(GS_EXE) and not os.path.exists(GS_EXE):
                raise HTTPException(500, f"Ghostscript not found at: {GS_EXE}")
            pages_text = ocr_pdf(tmp_path, password)
            total_pages = len(pages_text)
            blocks = extract_cc_blocks(pages_text)
            all_cc = [parse_cc_lines(b) for b in blocks]
            cc_combined = pd.concat([d for d in all_cc if not d.empty], ignore_index=True) \
                          if any(not d.empty for d in all_cc) else pd.DataFrame()
            df_out = format_cc_output(cc_combined)
            detail_total = float(
                pd.to_numeric(df_out.get("Debit"), errors="coerce").fillna(0).sum()
            )
        # ── AIF ───────────────────────────────────────────────────────────
        elif mode == "aif":
            df_out, problems, total_pages, detail_total, summary_total, ext_method = parse_aif_pdf(tmp_path, password)
            base_rows = max(len(df_out) - 1, 0)
            if ext_method == "text" and base_rows < MIN_ROWS:
                df_retry, problems_retry, pages_retry, detail_retry, summary_retry, method_retry = parse_aif_pdf(
                    tmp_path,
                    password,
                    force_ocr=True,
                )
                retry_base_rows = max(len(df_retry) - 1, 0)
                if retry_base_rows > base_rows:
                    df_out, problems, total_pages, detail_total, summary_total, ext_method = (
                        df_retry,
                        problems_retry,
                        pages_retry,
                        detail_retry,
                        summary_retry,
                        method_retry,
                    )
            method = f"aif_{ext_method}"

            if summary_total is not None:
                diff = detail_total - float(summary_total)
                if abs(diff) <= 0.01:
                    status, match_text = "MATCH", \
                        "Detail TOTAL matches Summary Net Contribution (within rounding tolerance)."
                else:
                    status, match_text = "MISMATCH", \
                        "Detail TOTAL does NOT match Summary Net Contribution."

        # ── AXIS CREDIT CARD ─────────────────────────────────────────────────
        elif mode == "axis":
            df_axis, problems, total_pages, ext_method = parse_axis_cc(tmp_path, password)
            if ext_method == "text" and len(df_axis) < MIN_ROWS:
                df_retry, problems_retry, pages_retry, method_retry = parse_axis_cc(
                    tmp_path,
                    password,
                    force_ocr=True,
                )
                if len(df_retry) > len(df_axis):
                    df_axis, problems, total_pages, ext_method = df_retry, problems_retry, pages_retry, method_retry
            df_out = add_axis_total_row(df_axis)
            detail_total = pd.to_numeric(df_axis["Amount"], errors="coerce").fillna(0).sum() if not df_axis.empty else 0
            method = f"axis_{ext_method}"
            status = "OK"
            match_text = f"Extracted {len(df_axis)} transactions. Net total: {detail_total:.2f}"

        # ── HDFC BANK ─────────────────────────────────────────────────────
        elif mode == "hdfc":
            df_hdfc, problems, total_pages, hdfc_meta, ext_method = parse_hdfc_bank(tmp_path, password)
            if ext_method == "text" and len(df_hdfc) < MIN_ROWS:
                df_retry, problems_retry, pages_retry, meta_retry, method_retry = parse_hdfc_bank(
                    tmp_path,
                    password,
                    force_ocr=True,
                )
                if len(df_retry) > len(df_hdfc):
                    df_hdfc, problems, total_pages, hdfc_meta, ext_method = (
                        df_retry,
                        problems_retry,
                        pages_retry,
                        meta_retry,
                        method_retry,
                    )
            df_out = add_hdfc_total_row(df_hdfc)
            detail_total = pd.to_numeric(df_hdfc["WithdrawalAmt"], errors="coerce").fillna(0).sum() if not df_hdfc.empty else 0
            method = f"hdfc_{ext_method}"

            # Verify against summary totals
            exp_dr = hdfc_meta.get("exp_dr_sum", 0)
            actual_dr = pd.to_numeric(df_hdfc["WithdrawalAmt"], errors="coerce").fillna(0).sum() if not df_hdfc.empty else 0
            actual_cr = pd.to_numeric(df_hdfc["DepositAmt"], errors="coerce").fillna(0).sum() if not df_hdfc.empty else 0
            exp_cr = hdfc_meta.get("exp_cr_sum", 0)

            dr_ok = abs(actual_dr - exp_dr) < 0.01
            cr_ok = abs(actual_cr - exp_cr) < 0.01
            if dr_ok and cr_ok:
                status = "MATCH"
                match_text = f"Withdrawal total ({actual_dr:,.2f}) and Deposit total ({actual_cr:,.2f}) match summary."
            else:
                status = "MISMATCH"
                parts = []
                if not dr_ok:
                    parts.append(f"Withdrawal: parsed {actual_dr:,.2f} vs expected {exp_dr:,.2f}")
                if not cr_ok:
                    parts.append(f"Deposit: parsed {actual_cr:,.2f} vs expected {exp_cr:,.2f}")
                match_text = "Mismatch — " + "; ".join(parts)

        # ── ICICI BANK ────────────────────────────────────────────────────
        elif mode == "icici":
            df_icici, problems, total_pages, ext_method = parse_icici_bank(tmp_path, password)
            if ext_method == "text" and len(df_icici) < MIN_ROWS:
                df_retry, problems_retry, pages_retry, method_retry = parse_icici_bank(
                    tmp_path,
                    password,
                    force_ocr=True,
                )
                if len(df_retry) > len(df_icici):
                    df_icici, problems, total_pages, ext_method = df_retry, problems_retry, pages_retry, method_retry
            df_out = add_icici_total_row(df_icici)
            detail_total = pd.to_numeric(df_icici["WITHDRAWALS"], errors="coerce").fillna(0).sum() if not df_icici.empty else 0
            method = f"icici_{ext_method}"

            # Compare deposits/withdrawals for verification
            actual_dep = pd.to_numeric(df_icici["DEPOSITS"], errors="coerce").fillna(0).sum() if not df_icici.empty else 0
            actual_wd = detail_total
            status = "OK"
            match_text = f"Extracted {len(df_icici)} transactions. Deposits: {actual_dep:,.2f}, Withdrawals: {actual_wd:,.2f}"

        else:
            raise HTTPException(400, f"Unknown mode/sub_mode: {mode}/{sub_mode}")

        # ── Serialize to Excel + XML ──────────────────────────────────────────
        if df_out.empty or len(df_out) < MIN_ROWS:
            raise HTTPException(422, f"Extraction produced fewer than {MIN_ROWS} rows. "
                                     "Check that this is the right statement type.")

        # ── Save output files to output folder ────────────────────────────────
        stem = Path(filename).stem
        xlsx_bytes = df_to_excel_bytes(df_out)
        xml_bytes  = df_to_xml_bytes(df_out)

        xlsx_name = f"{stem}_transactions.xlsx"
        xml_name  = f"{stem}_transactions.xml"

        (OUTPUT_DIR / xlsx_name).write_bytes(xlsx_bytes)
        (OUTPUT_DIR / xml_name).write_bytes(xml_bytes)

        xlsx_b64   = b64(xlsx_bytes)
        xml_b64    = b64(xml_bytes)
        rows_count = max(len(df_out) - (1 if mode in ("pms", "aif") else 0), 0)

        # Serialize rows as JSON (NaN/Inf -> None)
        import math
        rows_json = _safe_rows_json(df_out)

        # Sanitise totals too
        if isinstance(detail_total, float) and (math.isnan(detail_total) or math.isinf(detail_total)):
            detail_total = 0.0
        if isinstance(summary_total, float) and (math.isnan(summary_total) or math.isinf(summary_total)):
            summary_total = None

        return {
            "ok": True,
            "rows": rows_json,
            "report": {
                "pdfName": filename,
                "runTime": now_str,
                "pagesProcessed": total_pages,
                "rowsExtracted": rows_count,
                "problemLines": problems,
                "detailTotal": detail_total,
                "summaryTotal": summary_total,
                "status": status,
                "matchText": match_text,
                "method": method,
                
                "outputFiles": {
                    "excel": xlsx_name,
                    "xml":   xml_name,
                    "report": f"{stem}_extraction_report.txt",
                },
                "mode": "pms" if mode == "pms" else "aif" if mode == "aif" else "axis" if mode == "axis" else "hdfc" if mode == "hdfc" else "icici" if mode == "icici" else sub_mode,
            },
            "files": {
                "xlsx": xlsx_b64,
                "xml":  xml_b64,
            },
        }

    except HTTPException:
        raise
    except MemoryError:
        gc.collect()
        raise HTTPException(507, "Server ran out of memory processing this PDF. Try a smaller file.")
    except subprocess.TimeoutExpired as e:
        gc.collect()
        timeout_secs = int(getattr(e, "timeout", 0) or 0)
        raise HTTPException(
            504,
            f"OCR processing timed out after {timeout_secs} seconds. "
            "Try a smaller PDF or split large files into parts.",
        )
    except Exception as e:
        raise HTTPException(500, f"Extraction failed: {repr(e)}")


# =============================================================================
# ── /extract-text ENDPOINT ────────────────────────────────────────────────────
# =============================================================================

@app.post("/extract-text")
async def extract_text(
    request: Request,
    file: UploadFile = File(...),
    password: Optional[str] = Form(None),
):
    _get_current_user(request)  # require authentication
    """Extract raw text from each page of a PDF for the Custom Format Builder."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are accepted.")

    content = await file.read()
    safe_name = _sanitize_filename(file.filename)
    input_path = INPUT_DIR / safe_name
    input_path.write_bytes(content)
    tmp_path = str(input_path)

    try:
        pages_text, ext_method = extract_pdf_pages_text(tmp_path, password)
        pages = []
        for page_idx, text in enumerate(pages_text, 1):
            lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
            pages.append({"page": page_idx, "lines": lines})
        return JSONResponse({"ok": True, "pages": pages, "method": ext_method})
    except Exception as e:
        raise HTTPException(500, f"Text extraction failed: {repr(e)}")


# =============================================================================
# ── /custom-convert ENDPOINT ─────────────────────────────────────────────────
# =============================================================================

import json

@app.post("/custom-convert")
async def custom_convert(
    request: Request,
    file: UploadFile = File(...),
    config_json: str = Form(...),
    password: Optional[str] = Form(None),
):
    _get_current_user(request)  # require authentication
    """Parse a PDF using a user-defined custom format config with intelligent extraction."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are accepted.")

    try:
        config = json.loads(config_json)
    except json.JSONDecodeError:
        raise HTTPException(400, "Invalid config_json.")

    header_keyword = config.get("headerKeyword", "")
    end_keyword = config.get("endKeyword", "")
    col_defs = config.get("columns", [])
    if not col_defs:
        raise HTTPException(400, "No columns defined.")

    content = await file.read()
    safe_name = _sanitize_filename(file.filename)
    input_path = INPUT_DIR / safe_name
    input_path.write_bytes(content)
    tmp_path = str(input_path)

    col_names = [c["name"] for c in col_defs]
    col_types = [c.get("type", "text") for c in col_defs]
    num_cols = len(col_defs)

    # ── Identify which column indices are date / number / text ────────
    date_cols = [i for i, t in enumerate(col_types) if t == "date"]
    num_cols_idx = [i for i, t in enumerate(col_types) if t == "number"]
    text_cols = [i for i, t in enumerate(col_types) if t == "text"]

    # Build regex helpers
    CUSTOM_DATE_RE = re.compile(
        r"\b(\d{1,2})[\/\-\.\s](\d{1,2}|\w{3,9})[\/\-\.\s](\d{2,4})\b"
    )
    CUSTOM_AMOUNT_RE = re.compile(r"-?\d[\d,]*\.?\d*")
    CUSTOM_FOOTER_RE = re.compile(
        r"(total|summary|grand\s+total|closing\s+balance|statement\s+summary"
        r"|opening\s+balance|disclaimer|this\s+is\s+a\s+computer"
        r"|page\s+\d+\s+of\s+\d+|terms\s+and\s+conditions|note\s*:"
        r"|e\s*&\s*o\.?\s*e|registered\s+office|www\.|http"
        r"|customer\s+care|contact\s+us|helpline)",
        re.IGNORECASE,
    )

    try:
        # ── Strategy 1: Try pdfplumber table extraction first ─────────
        rows = []
        open_kwargs = {"path_or_fp": tmp_path}
        if password:
            open_kwargs["password"] = password

        tables_found = False
        with pdfplumber.open(**open_kwargs) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables({"vertical_strategy": "text", "horizontal_strategy": "text"})
                for table in (tables or []):
                    if not table:
                        continue
                    for trow in table:
                        if not trow or all(not cell or not str(cell).strip() for cell in trow):
                            continue
                        cells = [clean(str(c)) if c else "" for c in trow]
                        joined = " ".join(cells)
                        # Skip header row itself
                        if header_keyword and header_keyword.lower() in joined.lower():
                            continue
                        # Skip footer/noise
                        if CUSTOM_FOOTER_RE.search(joined):
                            continue
                        # Must have at least one date or number to be a data row
                        has_date = bool(CUSTOM_DATE_RE.search(joined))
                        has_num = bool(CUSTOM_AMOUNT_RE.search(joined))
                        if not has_date and not has_num:
                            continue
                        tables_found = True
                        # Map cells to columns
                        row = _map_cells_to_columns(cells, col_names, col_types, num_cols)
                        if row:
                            rows.append(row)

        # ── Strategy 2: Fall back to text-based line parsing ──────────
        if not tables_found or len(rows) < 3:
            rows = []  # reset
            all_lines = []
            with pdfplumber.open(**open_kwargs) as pdf:
                for page in pdf.pages:
                    text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
                    for l in text.splitlines():
                        stripped = l.strip()
                        if stripped:
                            all_lines.append(stripped)

            # Find data region between header and end keywords
            capture = not bool(header_keyword)
            data_lines = []
            for line in all_lines:
                if not capture:
                    if header_keyword and header_keyword.lower() in line.lower():
                        capture = True
                    continue
                if end_keyword and end_keyword.lower() in line.lower():
                    break
                # Skip footer/noise lines
                if CUSTOM_FOOTER_RE.search(line):
                    continue
                if PAGE_FOOTER_RE.search(line):
                    continue
                data_lines.append(line)

            # Intelligent line parsing with multi-line description merging
            pending_row = None
            for line in data_lines:
                # Check if this line starts a new transaction (has a date)
                date_match = CUSTOM_DATE_RE.search(line)
                has_amount = bool(CUSTOM_AMOUNT_RE.search(line))

                if date_match and date_cols:
                    # Flush pending row
                    if pending_row and any(v is not None and str(v).strip() for v in pending_row.values()):
                        rows.append(pending_row)

                    # Parse this line into a new row
                    pending_row = _parse_line_intelligent(
                        line, col_names, col_types, num_cols,
                        CUSTOM_DATE_RE, CUSTOM_AMOUNT_RE
                    )
                elif not date_cols and has_amount:
                    # No date columns defined — each line with an amount is a row
                    if pending_row and any(v is not None and str(v).strip() for v in pending_row.values()):
                        rows.append(pending_row)
                    pending_row = _parse_line_intelligent(
                        line, col_names, col_types, num_cols,
                        CUSTOM_DATE_RE, CUSTOM_AMOUNT_RE
                    )
                elif pending_row and text_cols:
                    # Continuation line — merge into last text column
                    last_text_col = col_names[text_cols[-1]]
                    existing = pending_row.get(last_text_col, "") or ""
                    pending_row[last_text_col] = clean(existing + " " + line)
                else:
                    # Standalone line, try to parse it
                    row = _parse_line_intelligent(
                        line, col_names, col_types, num_cols,
                        CUSTOM_DATE_RE, CUSTOM_AMOUNT_RE
                    )
                    if row and any(v is not None and str(v).strip() for v in row.values()):
                        rows.append(row)

            # Flush last pending row
            if pending_row and any(v is not None and str(v).strip() for v in pending_row.values()):
                rows.append(pending_row)

        return JSONResponse({"ok": True, "rows": rows, "totalLines": len(rows), "parsedRows": len(rows)})

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Custom conversion failed: {repr(e)}")


def _map_cells_to_columns(cells, col_names, col_types, num_cols):
    """Map extracted table cells to column definitions."""
    row = {}
    # If cell count matches column count, direct mapping
    if len(cells) >= num_cols:
        for i in range(num_cols):
            val = cells[i] if i < len(cells) else ""
            if col_types[i] == "date":
                row[col_names[i]] = parse_any_date(val) or val
            elif col_types[i] == "number":
                row[col_names[i]] = clean_number(val)
            else:
                row[col_names[i]] = clean(val)
    else:
        # Fewer cells than columns — merge remaining into last text column
        for i in range(min(len(cells), num_cols)):
            val = cells[i]
            if col_types[i] == "date":
                row[col_names[i]] = parse_any_date(val) or val
            elif col_types[i] == "number":
                row[col_names[i]] = clean_number(val)
            else:
                row[col_names[i]] = clean(val)
        for i in range(len(cells), num_cols):
            row[col_names[i]] = None
    return row


def _parse_line_intelligent(line, col_names, col_types, num_cols, date_re, amount_re):
    """Parse a single text line into columns using type-aware extraction."""
    row = {name: None for name in col_names}
    remaining = line

    # Pass 1: Extract dates
    for i, (name, ctype) in enumerate(zip(col_names, col_types)):
        if ctype != "date":
            continue
        m = date_re.search(remaining)
        if m:
            raw_date = m.group(0)
            parsed = parse_any_date(raw_date)
            row[name] = parsed or raw_date
            remaining = remaining[:m.start()] + " " + remaining[m.end():]
            remaining = clean(remaining)

    # Pass 2: Extract numbers (from the right side, as amounts typically are rightmost)
    number_cols_reversed = [(i, col_names[i]) for i in range(num_cols) if col_types[i] == "number"]
    number_cols_reversed.reverse()
    amounts_found = list(amount_re.finditer(remaining))

    for (ci, cname), amt_match in zip(number_cols_reversed, reversed(amounts_found)):
        val = clean_number(amt_match.group(0))
        if val is not None:
            row[cname] = val
            remaining = remaining[:amt_match.start()] + " " + remaining[amt_match.end():]
            remaining = clean(remaining)

    # Check for Cr/Credit indicator for debit/credit split
    is_credit = bool(re.search(r"\bCr\b|\bCR\b|\bcredit\b", remaining, re.IGNORECASE))
    remaining = re.sub(r"\b(Cr|CR|Dr|DR)\b", "", remaining).strip()
    remaining = clean(remaining)

    # If there's a "Debit" and "Credit" column, assign based on indicator
    debit_idx = next((i for i, n in enumerate(col_names) if "debit" in n.lower()), None)
    credit_idx = next((i for i, n in enumerate(col_names) if "credit" in n.lower()), None)
    if debit_idx is not None and credit_idx is not None:
        # Find the single amount that was extracted for whichever number col
        for i in range(num_cols):
            if col_types[i] == "number" and row[col_names[i]] is not None:
                amt = row[col_names[i]]
                if i == debit_idx or i == credit_idx:
                    continue  # already correctly assigned
                if is_credit:
                    row[col_names[credit_idx]] = amt
                    row[col_names[debit_idx]] = None
                else:
                    row[col_names[debit_idx]] = amt
                    row[col_names[credit_idx]] = None
                row[col_names[i]] = None
                break

    # Pass 3: Remaining text goes to text columns
    text_col_names = [col_names[i] for i in range(num_cols) if col_types[i] == "text"]
    if text_col_names and remaining.strip():
        # If multiple text columns, try to split by 2+ spaces
        if len(text_col_names) > 1:
            parts = re.split(r"\s{2,}", remaining.strip())
            for j, tcol in enumerate(text_col_names):
                if j < len(parts):
                    row[tcol] = clean(parts[j])
        else:
            row[text_col_names[0]] = clean(remaining)

    return row


@app.get("/download/{filename}")
def download_file(filename: str):
    """Download a file from the output folder."""
    file_path = OUTPUT_DIR / filename
    if not file_path.exists():
        raise HTTPException(404, f"File not found: {filename}")
    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type="application/octet-stream",
    )


# =============================================================================
# ── /run-custom-code ENDPOINT ─────────────────────────────────────────────────
# =============================================================================

import traceback as _tb
import math

ALLOWED_IMPORTS = {
    "re", "io", "json", "csv", "math", "datetime", "collections",
    "pdfplumber", "pandas", "pd",
}

@app.post("/run-custom-code")
async def run_custom_code(
    request: Request,
    file: UploadFile = File(...),
    code: str = Form(...),
    password: Optional[str] = Form(None),
):
    _get_current_user(request)  # require authentication
    """
    Execute user-provided Python parser code against an uploaded PDF.
    
    The code MUST define a function:
        def parse(pdf_path: str, password: str | None) -> list[dict]
    
    Available in scope: pdfplumber, pandas (as pd), re, io, json, csv, math,
    datetime, collections, clean(), clean_number(), parse_any_date().
    """
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are accepted.")

    if not code or not code.strip():
        raise HTTPException(400, "No code provided.")

    # Save file
    content = await file.read()
    safe_name = _sanitize_filename(file.filename)
    input_path = INPUT_DIR / safe_name
    input_path.write_bytes(content)
    tmp_path = str(input_path)

    # Build a safe-ish execution namespace
    exec_globals = {
        "__builtins__": {
            # Safe builtins only
            "print": print, "len": len, "range": range, "enumerate": enumerate,
            "zip": zip, "map": map, "filter": filter, "sorted": sorted,
            "reversed": reversed, "list": list, "dict": dict, "set": set,
            "tuple": tuple, "str": str, "int": int, "float": float, "bool": bool,
            "None": None, "True": True, "False": False,
            "isinstance": isinstance, "type": type, "hasattr": hasattr,
            "getattr": getattr, "abs": abs, "round": round, "sum": sum,
            "min": min, "max": max, "any": any, "all": all,
            "ValueError": ValueError, "TypeError": TypeError,
            "KeyError": KeyError, "IndexError": IndexError,
            "Exception": Exception, "RuntimeError": RuntimeError,
            "StopIteration": StopIteration,
            "next": next, "iter": iter,
        },
        # Pre-imported modules
        "re": re,
        "io": io,
        # "os" removed — security risk (shell access)
        "json": json,
        "math": math,
        "pdfplumber": pdfplumber,
        "pd": pd,
        "pandas": pd,
        "datetime": datetime,
        # Helper functions from StatementIQ
        "clean": clean,
        "clean_number": clean_number,
        "parse_any_date": parse_any_date,
    }

    try:
        import collections
        exec_globals["collections"] = collections
    except ImportError:
        pass

    try:
        import csv
        exec_globals["csv"] = csv
    except ImportError:
        pass

    try:
        # Execute the user code to define the parse() function
        exec(code, exec_globals)

        if "parse" not in exec_globals or not callable(exec_globals["parse"]):
            raise HTTPException(400,
                "Your code must define a function: def parse(pdf_path: str, password: str | None) -> list[dict]")

        # Call the parse function
        result = exec_globals["parse"](tmp_path, password if password else None)

        if not isinstance(result, list):
            raise HTTPException(422, f"parse() must return a list[dict], got {type(result).__name__}")

        # Sanitise output
        rows = []
        for row in result:
            if not isinstance(row, dict):
                continue
            sanitised = {}
            for k, v in row.items():
                if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                    v = None
                sanitised[str(k)] = v
            rows.append(sanitised)

        # Get column names from first row
        col_names = list(rows[0].keys()) if rows else []

        return JSONResponse({
            "ok": True,
            "rows": rows,
            "columns": col_names,
            "totalRows": len(rows),
        })

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.exception("run-custom-code error")
        # Return only a sanitized error message, not the full traceback
        error_msg = str(e)
        if os.environ.get("DEBUG", "").lower() == "true":
            error_msg += "\n\n" + _tb.format_exc()
        return JSONResponse({
            "ok": False,
            "error": f"Parsing failed: {error_msg}",
        }, status_code=422)


# =============================================================================
# ── AIF TRANSACTION PARSER ────────────────────────────────────────────────────
# =============================================================================

AIF_DATE_RE = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$")
TXN_DETAILS_RE = re.compile(r"\bTransaction\s+Details\b", re.IGNORECASE)
AIF_SUMMARY_STOP_RE = re.compile(r"^\s*Summary\s+as\s+on\b", re.IGNORECASE)
NET_CONTRIB_HDR_RE = re.compile(r"\bNet\s+Contribution\s*\(INR\)\b", re.IGNORECASE)
MONEY_TOKEN_RE = re.compile(r"^\(?-?[\d,]+(?:\.\d+)?\)?$")

AIF_COLS = [
    "Transaction Date", "Transaction Description", "Class",
    "Distribution (INR)", "Gross Contribution (INR)", "Post tax NAV (INR)",
    "Units Allotted", "Net Contribution",
]


def _aif_is_date(tok):
    return bool(AIF_DATE_RE.match(tok))


def _aif_clean_money(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in {"", "-", "—"}:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()
    s = s.replace(",", "")
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def _aif_has_class(tokens):
    return any(
        t.upper() == "CLASS" or (t.upper().startswith("CLASS") and t.upper() != "CLASS")
        for t in tokens
    )


def _aif_is_header_or_title_line(line):
    toks = line.split()
    if toks and _aif_is_date(toks[0]):
        return False
    s = " ".join(line.split()).strip().lower()
    if not s:
        return False
    has_inr = "(inr)" in s
    has_date_word = "date" in s
    has_contrib = "contribution" in s
    has_nav = "nav" in s
    has_units = "units" in s
    has_gross = "gross" in s
    if has_inr and (has_date_word or has_contrib or has_gross or has_nav or has_units):
        return True
    header_phrases = [
        "transaction details", "transaction date", "transaction description",
        "distribution (inr)", "gross contribution", "post tax nav",
        "units allotted", "net contribution", "class",
    ]
    hits = sum(1 for h in header_phrases if h in s)
    if hits >= 2:
        return True
    if s.endswith("fund") and not any(ch.isdigit() for ch in s):
        return True
    return False


def _aif_is_desc_only_line(line):
    if AIF_SUMMARY_STOP_RE.search(line):
        return False
    if _aif_is_header_or_title_line(line):
        return False
    toks = line.split()
    if not toks:
        return False
    if _aif_is_date(toks[0]):
        return False
    if _aif_has_class(toks):
        return False
    if any(MONEY_TOKEN_RE.match(t) for t in toks):
        return False
    return True


def _aif_extract_summary_net_contribution(lines):
    for i, ln in enumerate(lines):
        if NET_CONTRIB_HDR_RE.search(ln):
            for j in range(i + 1, min(i + 8, len(lines))):
                row = lines[j].strip()
                if not row:
                    continue
                toks = row.split()
                money = [t for t in toks if MONEY_TOKEN_RE.match(t)]
                if money:
                    return _aif_clean_money(money[-1])
            return None
    return None


def _aif_try_parse_txn_row(tokens):
    if len(tokens) < 6:
        return None
    if not _aif_is_date(tokens[0]):
        return None
    class_idx = None
    class_text = None
    for j in range(1, len(tokens) - 1):
        if tokens[j].upper() == "CLASS":
            class_idx = j
            class_text = f"{tokens[j]} {tokens[j+1]}"
            break
        if tokens[j].upper().startswith("CLASS") and tokens[j].upper() != "CLASS":
            class_idx = j
            class_text = tokens[j]
            break
    if class_idx is None:
        return None
    date = tokens[0]
    desc = " ".join(tokens[1:class_idx]).strip()
    tail_tokens = tokens[class_idx + 1:]
    money_like = []
    for t in tail_tokens:
        if t in {"-", "—"} or MONEY_TOKEN_RE.match(t):
            money_like.append(t)
    if len(money_like) < 5:
        return None
    dist_s, gross_s, nav_s, units_s, net_s = money_like[-5:]
    return {
        "Transaction Date": date,
        "Transaction Description": desc,
        "Class": class_text.strip(),
        "Distribution (INR)": _aif_clean_money(dist_s),
        "Gross Contribution (INR)": _aif_clean_money(gross_s),
        "Post tax NAV (INR)": _aif_clean_money(nav_s),
        "Units Allotted": _aif_clean_money(units_s),
        "Net Contribution": _aif_clean_money(net_s),
    }


def _aif_extract_transactions(lines):
    in_txn = False
    rows = []
    problems = []
    next_desc_buffer = []
    i = 0
    while i < len(lines):
        ln = lines[i].strip()
        if _aif_is_header_or_title_line(ln):
            next_desc_buffer = []
            i += 1
            continue
        if TXN_DETAILS_RE.search(ln):
            in_txn = True
            next_desc_buffer = []
            i += 1
            continue
        if not in_txn:
            i += 1
            continue
        if AIF_SUMMARY_STOP_RE.search(ln):
            break
        if not ln:
            i += 1
            continue
        toks = ln.split()
        if _aif_is_desc_only_line(ln):
            words = ln.split()
            if len(words) <= 3 and rows:
                rows[-1]["Transaction Description"] += " " + ln
            else:
                next_desc_buffer.append(ln)
            i += 1
            continue
        if toks and _aif_is_date(toks[0]):
            combined = toks[:]
            if next_desc_buffer:
                combined = [combined[0]] + " ".join(next_desc_buffer).split() + combined[1:]
                next_desc_buffer = []
            parsed = _aif_try_parse_txn_row(combined)
            join_tries = 0
            while parsed is None and (i + 1) < len(lines) and join_tries < 12:
                nxt = lines[i + 1].strip()
                if not nxt:
                    i += 1
                    join_tries += 1
                    continue
                if _aif_is_header_or_title_line(nxt) or _aif_is_desc_only_line(nxt):
                    break
                nxt_toks = nxt.split()
                if nxt_toks and _aif_is_date(nxt_toks[0]):
                    break
                combined.extend(nxt_toks)
                parsed = _aif_try_parse_txn_row(combined)
                i += 1
                join_tries += 1
            if parsed is None:
                problems.append({"reason": "Could not parse date-start row", "line": " ".join(combined)})
            else:
                rows.append(parsed)
            i += 1
            continue
        i += 1
    df = pd.DataFrame(rows)
    if not df.empty:
        for c in AIF_COLS:
            if c not in df.columns:
                df[c] = None
        df = df[AIF_COLS]
        for c in ["Distribution (INR)", "Gross Contribution (INR)", "Post tax NAV (INR)", "Units Allotted", "Net Contribution"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df, problems


def _aif_add_total_row(df):
    if df is None or df.empty:
        df = pd.DataFrame(columns=AIF_COLS)
        total = 0.0
    else:
        total = float(pd.to_numeric(df["Net Contribution"], errors="coerce").fillna(0).sum())
    total = round(total, 2)
    total_row = {c: "" for c in AIF_COLS}
    total_row["Transaction Description"] = "TOTAL"
    total_row["Net Contribution"] = total
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True), total


def parse_aif_pdf(pdf_path: str, password: Optional[str] = None, force_ocr: bool = False):
    """Parse AIF statement PDF and return (df_with_total, problems, total_pages, detail_total, summary_total, extraction_method)."""
    # Use OCR-fallback text extraction
    pages_text, extraction_method = extract_pdf_pages_text(pdf_path, password, force_ocr=force_ocr)
    total_pages = len(pages_text)

    lines = []
    for text in pages_text:
        for ln in (text or "").splitlines():
            ln2 = ln.strip()
            if ln2:
                lines.append(ln2)

    summary_total = _aif_extract_summary_net_contribution(lines)
    df, problems = _aif_extract_transactions(lines)
    df_with_total, detail_total = _aif_add_total_row(df)
    return df_with_total, problems, total_pages, detail_total, summary_total, extraction_method


# =============================================================================
# ── /custom-format-request ENDPOINT ───────────────────────────────────────────
# =============================================================================

CUSTOM_FORMAT_DIR = Path(__file__).parent / "custom_format"
CUSTOM_FORMAT_DIR.mkdir(exist_ok=True)

@app.post("/custom-format-request")
async def custom_format_request(
    request: Request,
    file: UploadFile = File(...),
    password: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
):
    _require_admin(request)  # require admin role
    """Save a sample PDF for a custom format request."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are accepted.")

    content = await file.read()
    # Save with timestamp prefix to avoid overwriting
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^\w.\-]", "_", file.filename)
    dest = CUSTOM_FORMAT_DIR / f"{ts}_{safe_name}"
    dest.write_bytes(content)

    # Save metadata alongside
    meta = {
        "original_name": file.filename,
        "saved_as": dest.name,
        "timestamp": ts,
        "password": password or "",
        "notes": notes or "",
        "size_bytes": len(content),
    }
    meta_path = CUSTOM_FORMAT_DIR / f"{ts}_{safe_name}.json"
    meta_path.write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")

    return JSONResponse({"ok": True, "message": "Request saved", "file": dest.name, "request_id": f"{ts}_{safe_name}"})


@app.get("/custom-format-requests")
async def list_custom_format_requests(request: Request):
    """List all custom format requests with their metadata and admin replies. Admin only."""
    _require_admin(request)  # require admin role
    requests_list = []
    for meta_file in sorted(CUSTOM_FORMAT_DIR.glob("*.json"), reverse=True):
        try:
            meta = json.loads(meta_file.read_text(encoding="utf-8"))
            meta["request_id"] = meta_file.stem
            # Redact sensitive fields
            meta.pop("password", None)
            requests_list.append(meta)
        except Exception:
            continue
    return JSONResponse({"requests": requests_list})


@app.put("/custom-format-requests/{request_id}/reply")
async def reply_to_custom_format_request(
    request: Request,
    request_id: str,
    status: str = Form(...),       # "available", "processing", "uploaded", "rejected", "too_large"
    reply: str = Form(""),
    matched_format: Optional[str] = Form(None),  # e.g. "kotak" if already available
):
    """Admin replies to a custom format request with a status update."""
    _require_admin(request)  # require admin role
    meta_path = CUSTOM_FORMAT_DIR / f"{request_id}.json"
    if not meta_path.exists():
        raise HTTPException(404, "Request not found")

    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    meta["status"] = status
    meta["admin_reply"] = reply
    meta["matched_format"] = matched_format or ""
    meta["replied_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta_path.write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")

    return JSONResponse({"ok": True, "message": "Reply saved"})


@app.get("/health")
def health():
    return {"status": "ok", "ocr_available": OCR_AVAILABLE,
            "ghostscript": bool(shutil.which(GS_EXE) or os.path.exists(GS_EXE)),
            "ghostscript_path": shutil.which(GS_EXE) or GS_EXE,
            "tesseract": bool(shutil.which(TESSERACT_EXE) or os.path.exists(TESSERACT_EXE)) if OCR_AVAILABLE else False,
            "tesseract_path": shutil.which(TESSERACT_EXE) or TESSERACT_EXE if OCR_AVAILABLE else None}
